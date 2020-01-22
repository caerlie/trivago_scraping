from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from tkinter import messagebox
from bs4 import BeautifulSoup
import time, datetime
import openpyxl as px

version = "1.3.00"
print(version)

firstresearch = True
pagenum = 0
filename = datetime.datetime.now().strftime('%Y%m%d%H%M')

def checknext(c):
    try:
        c.find_element_by_class_name('btn--next')
        return True
    except:
        return False

def inittilt(nowsheet):
    nowsheet["A1"] = "Page-No"
    nowsheet.column_dimensions['A'].width = 15
    nowsheet["B1"] = "City"
    nowsheet.column_dimensions['B'].width = 15
    nowsheet["C1"] = "Hotelname"
    nowsheet.column_dimensions['C'].width = 50

    nowsheet["D1"] = "Bestbooking"
    nowsheet.column_dimensions['D'].width = 15
    nowsheet["E1"] = "Bestprice"
    nowsheet.column_dimensions['E'].width = 10
    nowsheet["F1"] = "Star"
    nowsheet.column_dimensions['F'].width = 10

    nowsheet["G1"] = "Other1"
    nowsheet.column_dimensions['G'].width = 15
    nowsheet["H1"] = "Other1"
    nowsheet.column_dimensions['H'].width = 10
    
    nowsheet["I1"] = "Other2"
    nowsheet.column_dimensions['I'].width = 15
    nowsheet["J1"] = "Other2"
    nowsheet.column_dimensions['J'].width = 10
    
    nowsheet["K1"] = "Other3"
    nowsheet.column_dimensions['K'].width = 15
    nowsheet["L1"] = "Other3"
    nowsheet.column_dimensions['L'].width = 10

    nowsheet["M1"] = "Checkin"
    nowsheet.column_dimensions['M'].width = 15
    nowsheet["N1"] = "Checkout"
    nowsheet.column_dimensions['N'].width = 15

    nowsheet["O1"] = "Currency"
    nowsheet.column_dimensions['O'].width = 10

def scraping(c, search_cityname, search_checkin, search_checkout, real_currency):
    global i
    global nowsheet
    global pagenum
    n = 1
    s = BeautifulSoup(c.page_source, "html.parser")
    hotelitems =  s.find_all("li", attrs={"class", "hotel-item item-order__list-item js_co_item"})
    pagenum = pagenum + 1

    dic_mouth = ['G', 'H', 'I', 'J', 'K', 'L']

    for hotelitem in hotelitems:
        try:
            '''
            星级信息筛选
            '''
            stardiv = hotelitem.find("div", attrs={"class", "stars-wrp"})
            if stardiv != None:
                stardiv = stardiv.find("meta")['content']
            else:
                stardiv = "0"
            if stardiv not in search_star:
                continue

            i+=1
            nowsheet["A" + str(i)] = str(pagenum) + '-' + str(n)

            nowsheet["B" + str(i)] = search_cityname

            hotelname = hotelitem.find("span", attrs={"class", "item-link name__copytext"}).text
            print(hotelname)
            nowsheet["C" + str(i)] = hotelname

            bookingsite = hotelitem.find("em", attrs={"class", "item__deal-best-ota"}).text
            print(bookingsite)
            nowsheet["D" + str(i)] = bookingsite

            hotelprice = hotelitem.find("strong", attrs={"class", "item__best-price"}).text
            print(hotelprice)
            try:
                nowsheet["E" + str(i)] = int(hotelprice.lstrip("$").replace(",",""))
            except:
                nowsheet["E" + str(i)] = hotelprice.lstrip("$").replace(",","")

            stardiv = hotelitem.find("div", attrs={"class", "stars-wrp"})
            if stardiv != None:
                star = stardiv.find("meta")['content']
                print(star)
                nowsheet["F" + str(i)] = star
            
            dealotheradvertisers = hotelitem.find_all("li", attrs={"class", "deal-other__offer js_co_deal"})
            dealnum = 0
            for dealotheradvertiser in dealotheradvertisers:
                if dealotheradvertiser.find("span") == None:
                    break
                nowsheet[dic_mouth[dealnum] + str(i)] = dealotheradvertiser.find("span").text
                dealnum = dealnum + 1
                
                try:
                    nowsheet[dic_mouth[dealnum] + str(i)] = int(dealotheradvertiser.find("strong").text.lstrip("$").replace(",",""))
                except:
                    nowsheet[dic_mouth[dealnum] + str(i)] = dealotheradvertiser.find("strong").text.lstrip("$").replace(",","")

                dealnum = dealnum + 1
            
            nowsheet["M" + str(i)] = search_checkin
            nowsheet["N" + str(i)] = search_checkout

            nowsheet["O" + str(i)] = real_currency

            n = n + 1
        except:
            i = i - 1
            continue

def clicknext(c):
    c.find_element_by_class_name('btn--next').click()
    WebDriverWait(c, 15).until(EC.presence_of_all_elements_located)
    time.sleep(8)

def inputcityname(c, cityname):
    '''
    输入城市名
    '''
    c.find_element_by_id('querytext').clear()
    time.sleep(1)
    c.find_element_by_id('querytext').send_keys(cityname)
    time.sleep(1)
    c.find_element_by_id('querytext').send_keys(Keys.ENTER)
    time.sleep(1)
    
def inputroomtype(chrome, search_roomtype):
    '''
    输入房间类型
    '''
    typeitems = chrome.find_elements_by_class_name("roomtype-btn__label")
    if len(typeitems) == 0 :
        chrome.find_elements_by_class_name("dealform-button--guests")[0].click()
        typeitems = chrome.find_elements_by_class_name("roomtype-btn__label")

    for typeitem in typeitems:
        if typeitem.text == search_roomtype:
            action = ActionChains(chrome)
            action.move_to_element_with_offset(typeitem, 5, 5)
            action.click()
            action.perform()
            time.sleep(1)
            return
    
    try:
        chrome.find_element_by_xpath("//input[@id='adults-input-2']").send_keys(Keys.BACK_SPACE)
        if search_roomtype == "Single room": 
            chrome.find_element_by_xpath("//input[@id='adults-input-2']").send_keys("1")
            chrome.find_element_by_class_name("btn--apply-config").click()
            time.sleep(1)
#            chrome.find_element_by_class_name("search-button").click()
            return
        elif search_roomtype == "Double room": 
            chrome.find_element_by_xpath("//input[@id='adults-input-2']").send_keys("2")
            chrome.find_element_by_class_name("btn--apply-config").click()
            time.sleep(1)
#            chrome.find_element_by_class_name("search-button").click()
            return
        else:
            chrome.find_element_by_xpath("//input[@id='adults-input-2']").send_keys("1")
            chrome.find_element_by_class_name("btn--apply-config").click()
            time.sleep(1)
#            chrome.find_element_by_class_name("search-button").click()
            return
    except:
#        chrome.find_element_by_class_name("search-button").click()
        return
            

def inputcurrency(chrome, search_currency):
    '''
    输入货币类型
    '''
    if search_currency == "HKD":
        return

    currency_element  = chrome.find_element_by_id('currency')
    currency_select_element = Select(currency_element)
    currency_select_element.select_by_value(search_currency)
    time.sleep(1)


def inputcal(chrome, search_date):
    '''
    取到打开日历的 年 月
    '''
    dic_mouth = {
        "January" : 1, 
        "February" : 2,
        "March" : 3,
        "April" : 4,
        "May" : 5,
        "June" : 6,
        "July" : 7,
        "August" : 8,
        "September" : 9,
        "October" : 10,
        "November" : 11, 
        "December" : 12
        }

    '''
    对比年 按前后箭头
    '''
    calheading_year =  int(chrome.find_element_by_class_name('cal-heading-month').find_elements_by_tag_name("span")[0].text.split()[0])
    search_year = int(search_date.split("-")[0])

    while search_year != calheading_year:
        if search_year > calheading_year:
            chrome.find_element_by_class_name('cal-btn-next').click()
            time.sleep(1)
        else:
            chrome.find_element_by_class_name('cal-btn-prev').click()
            time.sleep(1)

        calheading_year =  int(chrome.find_element_by_class_name('cal-heading-month').find_elements_by_tag_name("span")[0].text.split()[0])

    '''
    对比月 按前后箭头
    '''
    calheading_mouth =  dic_mouth[chrome.find_element_by_class_name('cal-heading-month').find_elements_by_tag_name("span")[0].text.split()[1]]
    search_mouth = int(search_date.split("-")[1])

    while search_mouth != calheading_mouth:
        if search_mouth > calheading_mouth:
            chrome.find_element_by_class_name('cal-btn-next').click()
            time.sleep(1)
        else:
            chrome.find_element_by_class_name('cal-btn-prev').click()
            time.sleep(1)
        
        calheading_mouth =  dic_mouth[chrome.find_element_by_class_name('cal-heading-month').find_elements_by_tag_name("span")[0].text.split()[1]]

    '''
    选取日
    '''
    action = ActionChains(chrome)
    action.move_to_element_with_offset(chrome.find_element_by_xpath("//time[@datetime='" + search_date + "']"), 5, 5)
    action.click()
    action.perform()
    time.sleep(1)

'''
取searchlist数据
'''
searchlist_exl = px.load_workbook('searchlist.xlsx')
search_sheet = searchlist_exl.active

'''
保存结果的excel生成
'''
retexl = px.Workbook()
nowsheet = retexl.active
i = 1

for row_num in range(2, search_sheet.max_row + 1):
    '''
    读取数据
    '''
    search_cityname = search_sheet["A" + str(row_num)].value
    if search_cityname == None:
        break

    search_checkin = search_sheet["B" + str(row_num)].value.strftime("%Y-%m-%d")
    search_checkout = search_sheet["C" + str(row_num)].value.strftime("%Y-%m-%d")

    search_roomtype = search_sheet["D" + str(row_num)].value
    search_currency = search_sheet["E" + str(row_num)].value

    if search_sheet["F" + str(row_num)].value == None:
        search_star = ['3', '4', '5']
    else:
        search_star = search_sheet["F" + str(row_num)].value.split(",")

    inittilt(nowsheet)

    chrome = webdriver.Chrome("chromedriver.exe")

    chrome.execute_script("window.open('','_blank');")
    chrome.switch_to.window(chrome.window_handles[0])

    chrome.get("https://www.trivago.hk/en?sLanguageLocale=UK")
    chrome.maximize_window()
    time.sleep(1)
    inputcurrency(chrome, search_currency)
    time.sleep(3)

    try:
        '''
        输入检索关键字
        '''
        inputcityname(chrome, search_cityname)
        inputcal(chrome, search_checkin)
        inputcal(chrome, search_checkout)
        inputroomtype(chrome, search_roomtype)
    except:
        continue
    
	
    chrome.find_element_by_class_name("search-button").click()
    time.sleep(8)

    pagenum = 0
    real_currency = "unkonw"

    try:
        '''
        取得真实货币类型
        '''
        currency_element  = chrome.find_element_by_id('currency')
        currency_select_element = Select(currency_element)
        real_currency = currency_select_element.first_selected_option.text
    except:
        real_currency = "unkonw"

    scraping(chrome, search_cityname, search_checkin, search_checkout, real_currency)
    while checknext(chrome):
        clicknext(chrome)
        scraping(chrome, search_cityname, search_checkin, search_checkout, real_currency)
        
    retexl.save('result/' + filename + '.xlsx')
    chrome.quit()



print("Finished! Press Enter to close...")
input()

    